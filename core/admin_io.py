"""Utilitários de import/export para o admin Django.

Implementação própria (não usa django-import-export por causa de
incompatibilidade com Django 6 no log_action singular).

Funcionalidades:
- exportar_modelo_csv: gera CSV via HttpResponse
- exportar_modelo_xlsx: gera XLSX via openpyxl
- importar_modelo_csv: parseia CSV com auto-deteccao de encoding/delimitador
- importar_modelo_xlsx: le XLSX com openpyxl
"""

import csv
import io
import unicodedata
from datetime import datetime

import openpyxl
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone


# Campos exportaveis por modelo (whitelist - evita expor dados sensiveis)
EXPORT_FIELDS = {
    'Servidor': [
        'id', 'nome', 'numero_documento', 'tipo_funcionario', 'plantao',
        'setor', 'veiculo', 'ativo',
    ],
    'RegistroAcesso': [
        'id', 'servidor', 'operador', 'tipo_acesso', 'data_hora',
        'data_hora_saida', 'operador_saida', 'isv', 'veiculo', 'setor',
        'saida_pendente', 'status_alteracao', 'observacao', 'observacao_saida',
        'justificativa',
    ],
    'User': [
        'id', 'username', 'first_name', 'last_name', 'email',
        'is_active', 'is_staff', 'is_superuser', 'last_login', 'date_joined',
    ],
    'LogAuditoria': [
        'id', 'usuario', 'tipo_acao', 'modelo', 'objeto_id', 'detalhes', 'data_hora',
    ],
}


def _normalizar_chave(s):
    """Lowercase, sem acentos, sem espacos extras."""
    s = str(s or '').strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


# Aliases para mapear nomes de coluna flexiveis para campos do model
IMPORT_ALIASES = {
    'Servidor': {
        'nome': ['nome', 'name'],
        'numero_documento': ['documento', 'cpf', 'matricula', 'numero_documento', 'numero do documento'],
        'setor': ['setor', 'lotacao'],
        'veiculo': ['veiculo', 'placa', 'veículo'],
        'tipo_funcionario': ['tipo', 'tipo_funcionario'],
        'ativo': ['ativo', 'active'],
    },
    'RegistroAcesso': {
        'id': ['id'],
        'servidor': ['servidor', 'servidor_id', 'servidor_nome', 'servidor_documento'],
        'operador': ['operador', 'operador_id', 'operador_username'],
        'tipo_acesso': ['tipo_acesso', 'tipo'],
        'data_hora': ['data_hora', 'data_hora_entrada'],
        'data_hora_saida': ['data_hora_saida'],
        'operador_saida': ['operador_saida'],
        'isv': ['isv'],
        'veiculo': ['veiculo', 'placa'],
        'setor': ['setor', 'lotacao'],
        'saida_pendente': ['saida_pendente'],
        'status_alteracao': ['status_alteracao', 'status'],
        'observacao': ['observacao'],
        'observacao_saida': ['observacao_saida'],
        'justificativa': ['justificativa'],
    },
    'User': {
        'username': ['username', 'usuario', 'login'],
        'first_name': ['first_name', 'nome', 'primeiro_nome'],
        'last_name': ['last_name', 'sobrenome'],
        'email': ['email', 'e-mail'],
        'is_active': ['is_active', 'ativo'],
        'is_staff': ['is_staff', 'staff'],
        'is_superuser': ['is_superuser', 'superuser', 'super'],
    },
}


def _mapear_colunas(row_dict, aliases):
    """Mapeia chaves normalizadas para nomes canonicos."""
    norm = {_normalizar_chave(k): k for k in row_dict.keys()}
    out = {}
    for canonico, lista in aliases.items():
        for alias in lista:
            if _normalizar_chave(alias) in norm:
                out[canonico] = row_dict[norm[_normalizar_chave(alias)]]
                break
    return out


def _parsear_bool(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        return v.strip().lower() in ('1', 'true', 'sim', 's', 'yes', 'y', 'ativo', 'verdadeiro')
    return False


def _parsear_arquivo(arquivo):
    """Le CSV ou XLSX e retorna lista de dicts."""
    nome = arquivo.name.lower() if hasattr(arquivo, 'name') else ''
    if nome.endswith('.xlsx') or nome.endswith('.xlsm'):
        return _ler_xlsx(arquivo)
    return _ler_csv(arquivo)


def _ler_csv(arquivo):
    """Lê CSV com auto-deteccao de encoding e delimitador."""
    conteudo = arquivo.read()
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    texto = None
    for enc in encodings:
        try:
            texto = conteudo.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        texto = conteudo.decode('latin-1', errors='replace')

    # Detecta delimitador
    try:
        sample = texto[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        delim = dialect.delimiter
    except csv.Error:
        delim = ';'

    return list(csv.DictReader(io.StringIO(texto), delimiter=delim))


def _ler_xlsx(arquivo):
    """Lê XLSX e retorna lista de dicts."""
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or '').strip() for c in rows[0]]
    resultado = []
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        resultado.append({header[i]: row[i] for i in range(len(header)) if i < len(row)})
    wb.close()
    return resultado


def _valor_para_celula(v):
    """Converte valor para formato serializavel em CSV/XLSX."""
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'sim' if v else 'nao'
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    if hasattr(v, 'isoformat'):
        return v.isoformat()
    return str(v)


def exportar_csv(request, queryset, model_name):
    """Gera HttpResponse com CSV do queryset."""
    fields = EXPORT_FIELDS.get(model_name, [])
    if not fields:
        raise ValueError(f'Modelo {model_name} nao suportado para export')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename={model_name.lower()}_{ts}.csv'
    response.write('﻿')  # BOM para Excel
    writer = csv.writer(response)
    writer.writerow(fields)
    for obj in queryset:
        row = []
        for f in fields:
            v = obj
            for part in f.split('__'):
                v = getattr(v, part, None) if v is not None else None
                if callable(v):
                    v = v()
            row.append(_valor_para_celula(v))
        writer.writerow(row)
    return response


def exportar_xlsx(request, queryset, model_name):
    """Gera HttpResponse com XLSX do queryset."""
    fields = EXPORT_FIELDS.get(model_name, [])
    if not fields:
        raise ValueError(f'Modelo {model_name} nao suportado para export')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name

    # Header
    ws.append(fields)
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='0a2540')
        cell.alignment = Alignment(horizontal='center')

    for obj in queryset:
        row = []
        for f in fields:
            v = obj
            for part in f.split('__'):
                v = getattr(v, part, None) if v is not None else None
                if callable(v):
                    v = v()
            row.append(_valor_para_celula(v))
        ws.append(row)

    # Auto-largura
    for col_idx, _ in enumerate(fields, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename={model_name.lower()}_{ts}.xlsx'
    wb.save(response)
    return response


def importar_arquivo(arquivo, model_name, usuario_log):
    """
    Importa arquivo (CSV/XLSX) para o modelo.
    Retorna dict com {criados, atualizados, erros: [str]}.
    """
    from core.models import LogAuditoria, RegistroAcesso, Servidor, PerfilUsuario
    from django.contrib.auth.models import User
    from core.utils import extrair_plantao_do_setor

    rows = _parsear_arquivo(arquivo)
    aliases = IMPORT_ALIASES.get(model_name, {})
    criados = 0
    atualizados = 0
    erros = []

    for i, row in enumerate(rows, start=2):  # linha 1 = header
        try:
            dados = _mapear_colunas(row, aliases)
            if model_name == 'Servidor':
                _importar_servidor_linha(dados)
                criado = True
            elif model_name == 'User':
                _importar_user_linha(dados)
                criado = True
            elif model_name == 'RegistroAcesso':
                _importar_registro_linha(dados)
                criado = True
            else:
                erros.append(f'Linha {i}: modelo {model_name} nao suportado')
                continue
            if criado:
                criados += 1
        except Exception as e:
            erros.append(f'Linha {i}: {e}')

    if criados or atualizados:
        LogAuditoria.objects.create(
            usuario=usuario_log,
            tipo_acao='CRIACAO',
            modelo=model_name,
            objeto_id=None,
            detalhes=f'Importacao em massa via admin: {criados} processados, {len(erros)} erros',
        )

    return {'criados': criados, 'atualizados': atualizados, 'erros': erros}


def _importar_servidor_linha(dados):
    from core.models import Servidor
    from core.utils import extrair_plantao_do_setor

    nome = (dados.get('nome') or '').strip()
    doc = (dados.get('numero_documento') or '').strip()
    setor = (dados.get('setor') or '').strip().upper()
    if not nome or not doc or not setor:
        raise ValueError('campos obrigatorios: nome, numero_documento, setor')

    veiculo = (str(dados.get('veiculo') or '')).strip().upper() or None
    tipo = (dados.get('tipo_funcionario') or 'PLANTONISTA').strip().upper()
    if tipo not in ['PLANTONISTA', 'EXPEDIENTE', 'VISITANTE', 'TERCEIRIZADO']:
        tipo = 'PLANTONISTA'
    ativo = _parsear_bool(dados.get('ativo'))

    Servidor.objects.update_or_create(
        numero_documento=doc,
        defaults={
            'nome': nome.upper(),
            'setor': setor,
            'veiculo': veiculo,
            'tipo_funcionario': tipo,
            'plantao': extrair_plantao_do_setor(setor),
            'ativo': ativo,
        },
    )


def _importar_user_linha(dados):
    from django.contrib.auth.models import User
    from core.models import PerfilUsuario

    username = (dados.get('username') or '').strip()
    if not username:
        raise ValueError('username obrigatorio')

    first_name = (dados.get('first_name') or '').strip()
    last_name = (dados.get('last_name') or '').strip()
    email = (dados.get('email') or '').strip()
    is_active = _parsear_bool(dados.get('is_active'))
    is_staff = _parsear_bool(dados.get('is_staff'))
    is_superuser = _parsear_bool(dados.get('is_superuser'))

    user, created = User.objects.update_or_create(
        username=username,
        defaults={
            'first_name': first_name,
            'last_name': last_name,
            'email': email,
            'is_active': is_active,
            'is_staff': is_staff,
            'is_superuser': is_superuser,
        },
    )
    if created:
        PerfilUsuario.objects.get_or_create(
            usuario=user,
            defaults={'tipo_usuario': 'OPERADOR'},
        )


def _importar_registro_linha(dados):
    from core.models import RegistroAcesso, Servidor
    from django.contrib.auth.models import User

    # Suporta ID direto ou busca por nome/documento
    reg_id = dados.get('id')
    servidor_input = dados.get('servidor')
    operador_input = dados.get('operador')

    servidor = None
    if reg_id:
        try:
            reg = RegistroAcesso.objects.get(pk=reg_id)
        except RegistroAcesso.DoesNotExist:
            raise ValueError(f'Registro id={reg_id} nao encontrado')
    else:
        # Criar novo registro basico
        if servidor_input:
            # Tenta por ID, depois por documento, depois por nome
            try:
                servidor = Servidor.objects.get(pk=int(servidor_input))
            except (ValueError, Servidor.DoesNotExist):
                try:
                    servidor = Servidor.objects.get(numero_documento=str(servidor_input))
                except Servidor.DoesNotExist:
                    servidor = Servidor.objects.filter(nome__icontains=str(servidor_input)).first()
            if not servidor:
                raise ValueError(f'Servidor "{servidor_input}" nao encontrado')

        operador = None
        if operador_input:
            try:
                operador = User.objects.get(pk=int(operador_input))
            except (ValueError, User.DoesNotExist):
                operador = User.objects.filter(username=str(operador_input)).first()
            if not operador:
                raise ValueError(f'Operador "{operador_input}" nao encontrado')

        if not servidor or not operador:
            raise ValueError('Para criar registro novo, informe servidor e operador')

        tipo = dados.get('tipo_acesso') or 'ENTRADA'
        if tipo not in ('ENTRADA', 'SAIDA'):
            tipo = 'ENTRADA'

        RegistroAcesso.objects.create(
            servidor=servidor,
            operador=operador,
            tipo_acesso=tipo,
            isv=_parsear_bool(dados.get('isv')),
            observacao=(dados.get('observacao') or '') or None,
        )